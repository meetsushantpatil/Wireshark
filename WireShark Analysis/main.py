'''
Created on Sep 9, 2016

@author: Sushant
'''
'''The project processes DNS packets captured from filtered packets, saved as .xlsx files & writes the DNS base name
on the column in front of it'''

#Open the file under process


from openpyxl import load_workbook
import read_folder

DNS_Table_File = load_workbook(filename='/Users/sushantpatil/Desktop/Project HawkEye/Public WiFi Project/Data/DNS Table_Priority.xlsx') 

#Read from input table

def create_DNS_Data_List(Keyword_column, Link_column,Row_start=4):
    #DNS_Table_File = load_workbook(filename='/Users/Sushant/Documents/Public WiFi Project/DNS Table_Priority.xlsx')

    DNS_Data_list = []

    print(DNS_Table_File.get_sheet_names())

    DNS_data = DNS_Table_File[DNS_Table_File.get_sheet_names()[0]];
    i = Row_start;
    while(DNS_data[Keyword_column + str(i)].value != None):
        a = DNS_data[Keyword_column + str(i)].value
        b = DNS_data[Link_column + str(i)].value
        i+=1;
        if((a!=None)):
            DNS_Data_list.append([a, b])
    return DNS_Data_list


#Match the strings
def match_DNS_Data(Read_Link_column,Write_Link_Column, DNS_Data_list,Row_start=2):
    #Input_Data_File = load_workbook(filename = '/Users/Sushant/Documents/Public WiFi Project/Data/Sushant_filtered_packets.xlsx')
    all_files = read_folder.get_all_excel_files_path("/Users/sushantpatil/Desktop/Project HawkEye/In Progress");
    for _file in all_files:
        print(_file);
        Input_Data_File = load_workbook(filename = _file);
        Input_Data = Input_Data_File[Input_Data_File.get_sheet_names()[0]]
    
        i = Row_start;
        while(Input_Data[Read_Link_column + str(i)].value != None):
            a = Input_Data[Read_Link_column + str(i)].value
            src_ip = Input_Data['C' + str(i)].value
            dst_ip = Input_Data['D' + str(i)].value
            if(src_ip !="192.168.1.1"):
                Input_Data["M" + str(i)].value = "User #"+str(src_ip[len(src_ip)-1]);
            if(dst_ip !="192.168.1.1"):
                Input_Data["M" + str(i)].value = "User #"+str(dst_ip[len(dst_ip)-1]);
            
            for DNS_entry in DNS_Data_list:
                if(a.find(DNS_entry[0])>=0):
                    Input_Data[Write_Link_Column + str(i)].value = DNS_entry[1]
            i+=1;
        Input_Data_File.save(_file[0:len(all_files[0])-5] + "_packets_Output.xlsx");
#Create the output file & save it

def main():
    DNS_Data_list = create_DNS_Data_List('B','C', 4)
    print(DNS_Data_list);
    match_DNS_Data('G','I',DNS_Data_list)
    
main()
    
#wb.save('/Users/Sushant/Documents/Public WiFi Project/Data/Sushant_filtered_packets.xlsx')
